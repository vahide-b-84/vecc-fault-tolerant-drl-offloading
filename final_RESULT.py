# final_RESULT.py
#   1) base            -> (bi-levels + flat) 
#   2) missing_data     -> (bi-level DQN)
#   3) trajectory_noise  -> (bi-level DQN)
#
#   <SCENARIO_TYPE>_results/<scenario>/

import os
import re
import math
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel

from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.data_source import NumDataSource, NumRef

from params import params


# -----------------------------
# General helpers
# -----------------------------
def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def _safe_float_from_folder_token(token: str) -> float:
    """
    token :
      "0_02" -> 0.02
      "0_20" -> 0.20
      "0_80" -> 0.80
      "1_00" -> 1.00
    """
    if token is None:
        return float("nan")
    s = token.replace("_", ".")
    try:
        return float(s)
    except Exception:
        return float("nan")


def _list_subfolders(root: str):
    if not os.path.exists(root):
        return []
    return sorted([d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))])


def _read_global_base_metrics(global_xlsx_path: str):

    logs = pd.read_excel(global_xlsx_path, sheet_name="Logs")
    needed_logs = ["Episode", "Avg Reward", "Avg Delay"]
    logs = logs[needed_logs].copy()

    summ = pd.read_excel(global_xlsx_path, sheet_name="Summary")
    if "normal_AVG_Failure" in summ.columns:
        failure_col = "normal_AVG_Failure"
    elif "AVG_Failure" in summ.columns:
        failure_col = "AVG_Failure"
    elif "FailureRate" in summ.columns:
        failure_col = "FailureRate"
    else:
        # 
        failure_col = summ.columns[min(4, len(summ.columns) - 1)]

    summ = summ[[failure_col]].copy()
    summ = summ.rename(columns={failure_col: "normal_AVG_Failure"})
    # concat by index 
    out = pd.concat([logs.reset_index(drop=True), summ.reset_index(drop=True)], axis=1)
    # out: Episode, Avg Reward, Avg Delay, normal_AVG_Failure
    return out


def _read_missing_data_metrics(global_xlsx_path: str):
    """
    missing_data:
      Logs: Episode, Avg Reward
      Summary: normal_AVG_Failure
    """
    logs = pd.read_excel(global_xlsx_path, sheet_name="Logs")
    logs = logs[["Episode", "Avg Reward"]].copy()

    summ = pd.read_excel(global_xlsx_path, sheet_name="Summary")
    if "normal_AVG_Failure" not in summ.columns:
        raise KeyError(f"'normal_AVG_Failure' not found in Summary: {global_xlsx_path}")
    summ = summ[["normal_AVG_Failure"]].copy()

    out = pd.concat([logs.reset_index(drop=True), summ.reset_index(drop=True)], axis=1)
    return out


def _read_params_p(global_xlsx_path: str, p_field: str):
    """
    از شیت Params مقدار missing_data_p یا trajectory_noise_p را اگر موجود باشد بخوان.
    """
    try:
        dfp = pd.read_excel(global_xlsx_path, sheet_name="Params")
        # Parameter, Value
        if "Parameter" in dfp.columns and "Value" in dfp.columns:
            row = dfp.loc[dfp["Parameter"] == p_field]
            if len(row):
                return float(row["Value"].values[0])
    except Exception:
        pass
    return None


# -----------------------------
# Chart utilities (Line)
# -----------------------------
DISTINCT_COLORS = [
    # 
    "1F77B4",  # blue
    "FF7F0E",  # orange
    "2CA02C",  # green
    "D62728",  # red
    "9467BD",  # purple
    "8C564B",  # brown
    "E377C2",  # pink
    "7F7F7F",  # gray
    "BCBD22",  # olive
    "17BECF",  # cyan
]


def _write_df_to_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))


def _add_line_chart(
    ws,
    title: str,
    x_col: int,
    series_cols: list,
    max_row: int,
    anchor: str,
):
    chart = LineChart()
    chart.title = title
    chart.width = 20
    chart.height = 8
    chart.legend.position = "r"
    chart.x_axis.title = "Episode"

    cats = Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
    chart.set_categories(cats)

    headers = [cell.value for cell in ws[1]]
    color_idx = 0

    for col_idx in series_cols:
        data_ref = Reference(ws, min_col=col_idx, max_col=col_idx, min_row=1, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        s = chart.series[-1]

        # legend label: 
        # 
        color = DISTINCT_COLORS[color_idx % len(DISTINCT_COLORS)]
        color_idx += 1
        s.graphicalProperties.line.solidFill = color

    ws.add_chart(chart, anchor)


# -----------------------------
# BASE mode: compare methods (Global + RSU sheets + RSU_avg)
# -----------------------------
def run_base_comparison(base_root: str, out_dir: str):
    """
    ساخت اکسل مقایسه روش‌های پوشه base
    """
    # 
    # dqn_flat: dqn_flat_0_00 
    base_methods = {
        "dqn": {"folder": "dqn_0_00", "suffix": "dqn", "display": "bi-level DQN"},
        "ppo": {"folder": "ppo_0_00", "suffix": "ppo", "display": "bi-level PPO"},
        "greedy": {"folder": "greedy_0_00", "suffix": "greedy", "display": "Greedy"},
        "NoForwarding": {"folder": "NoForwarding_0_00", "suffix": "NoForwarding", "display": "No-Forwarding"},
        "dqn_flat": {"folder": "dqn_flat_0_00", "suffix": "dqn", "display": "Flat DQN"},
    }

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Global_model sheet ----------
    global_data = {}
    for method_key, meta in base_methods.items():
        folder = os.path.join(base_root, meta["folder"])
        if not os.path.isdir(folder):
            continue

        global_xlsx = os.path.join(folder, f"Global_state_{params.FAILURE_STATE}_{meta['suffix']}.xlsx")
        if os.path.exists(global_xlsx):
            try:
                df = _read_global_base_metrics(global_xlsx)
                global_data[method_key] = df
            except Exception as e:
                print(f"[base] Error reading global {global_xlsx}: {e}")

    if global_data:
        _write_comp_sheet_with_charts(
            wb=wb,
            sheet_name="Global_model",
            data_dict=global_data,
            display_map={k: v["display"] for k, v in base_methods.items()},
            include_cols=("Avg Reward", "Avg Delay", "normal_AVG_Failure"),
        )
    else:
        print("[base] No global data found.")

    # ---------- Detect RSU IDs from any method folder ----------
    rsu_ids = set()
    for method_key, meta in base_methods.items():
        folder = os.path.join(base_root, meta["folder"])
        if not os.path.isdir(folder):
            continue
        for fn in os.listdir(folder):
            # 
            if fn.startswith("RSU_") and fn.endswith(f"_{meta['suffix']}.xlsx") and f"_state_{params.FAILURE_STATE}_" in fn:
                rsu_id = fn.split("_state_")[0]  # RSU_0
                rsu_ids.add(rsu_id)

    rsu_ids = sorted(rsu_ids)

    # ---------- RSU_avg collector ----------
    rsu_avg_collector = {k: [] for k in base_methods.keys()}

    # ---------- Per-RSU sheets ----------
    for rsu_id in rsu_ids:
        rsu_data = {}
        for method_key, meta in base_methods.items():
            folder = os.path.join(base_root, meta["folder"])
            if not os.path.isdir(folder):
                continue
            path = os.path.join(folder, f"{rsu_id}_state_{params.FAILURE_STATE}_{meta['suffix']}.xlsx")
            if os.path.exists(path):
                try:
                    df = _read_global_base_metrics(path)  # Logs+Summary
                    rsu_data[method_key] = df
                    rsu_avg_collector[method_key].append(df[["Episode", "Avg Reward", "Avg Delay", "normal_AVG_Failure"]])
                except Exception as e:
                    print(f"[base] Error reading RSU {path}: {e}")

        if rsu_data:
            _write_comp_sheet_with_charts(
                wb=wb,
                sheet_name=rsu_id,
                data_dict=rsu_data,
                display_map={k: v["display"] for k, v in base_methods.items()},
                include_cols=("Avg Reward", "Avg Delay", "normal_AVG_Failure"),
            )

    # ---------- RSU_avg ----------
    rsu_avg_data = {}
    for method_key, dfs in rsu_avg_collector.items():
        if dfs:
            avg_df = pd.concat(dfs, axis=0).groupby("Episode", as_index=False).mean(numeric_only=True)
            rsu_avg_data[method_key] = avg_df

    if rsu_avg_data:
        _write_comp_sheet_with_charts(
            wb=wb,
            sheet_name="RSU_avg",
            data_dict=rsu_avg_data,
            display_map={k: v["display"] for k, v in base_methods.items()},
            include_cols=("Avg Reward", "Avg Delay", "normal_AVG_Failure"),
        )

    out_path = os.path.join(out_dir, f"Final_Result_base_state_{params.FAILURE_STATE}.xlsx")
    wb.save(out_path)
    print(f"[base] Saved: {out_path}")


def _write_comp_sheet_with_charts(wb, sheet_name: str, data_dict: dict, display_map: dict, include_cols: tuple):
    """
    data_dict: {method_key: df(Episode, Avg Reward, Avg Delay, normal_AVG_Failure)}
    include_cols: ("Avg Reward","Avg Delay","normal_AVG_Failure")
    """
    ws = wb.create_sheet(sheet_name)

    merged = None
    for method_key, df in data_dict.items():
        df = df.copy()
        display_name = display_map.get(method_key, method_key)

        base_cols = ["Episode"] + list(include_cols)
        df = df[base_cols].copy()

        # rename with display suffix
        new_cols = ["Episode"]
        for c in include_cols:
            new_cols.append(f"{c}_{display_name}")
        df.columns = new_cols

        merged = df if merged is None else pd.merge(merged, df, on="Episode", how="outer")

    if merged is None or merged.empty:
        ws.append(["No data"])
        return

    # write df
    _write_df_to_sheet(ws, merged)

    max_row = ws.max_row
    headers = [cell.value for cell in ws[1]]
    ep_col = 1

    def find_series_cols(prefix: str):
        cols = []
        for idx, h in enumerate(headers, start=1):
            if isinstance(h, str) and h.startswith(prefix):
                cols.append(idx)
        return cols

    if any(isinstance(h, str) and h.startswith("Avg Reward_") for h in headers):
        _add_line_chart(
            ws,
            title="Avg Reward",
            x_col=ep_col,
            series_cols=find_series_cols("Avg Reward_"),
            max_row=max_row,
            anchor="L2",
        )

    if any(isinstance(h, str) and h.startswith("Avg Delay_") for h in headers):
        _add_line_chart(
            ws,
            title="Avg Delay",
            x_col=ep_col,
            series_cols=find_series_cols("Avg Delay_"),
            max_row=max_row,
            anchor="L20",
        )

    # failure
    if any(isinstance(h, str) and h.startswith("normal_AVG_Failure_") for h in headers):
        _add_line_chart(
            ws,
            title="Avg Task Failure Rate (%)",
            x_col=ep_col,
            series_cols=find_series_cols("normal_AVG_Failure_"),
            max_row=max_row,
            anchor="L38",
        )


# -----------------------------
# missing_data: line charts across all dqn_* folders (exclude dqn_flat)
# -----------------------------
def run_missing_data(root: str, out_dir: str):

    folders = _list_subfolders(root)
    dqn_folders = []
    for f in folders:
        if f.startswith("dqn_") and (not f.startswith("dqn_flat")):
            dqn_folders.append(f)

    if not dqn_folders:
        print("[missing_data] No dqn_* folders found.")
        return

    # collect per p
    series_reward = {}   # p_label -> df(Episode, Avg Reward)
    series_fail = {}     # p_label -> df(Episode, normal_AVG_Failure)

    for folder in sorted(dqn_folders):
        token = folder[len("dqn_"):]  # "0_02"
        p_float = _safe_float_from_folder_token(token)
        p_label = f"p={p_float:.2f}"

        global_xlsx = os.path.join(root, folder, f"Global_state_{params.FAILURE_STATE}_dqn.xlsx")
        if not os.path.exists(global_xlsx):
            print(f"[missing_data] Missing global file: {global_xlsx}")
            continue

        try:
            df = _read_missing_data_metrics(global_xlsx)
            series_reward[p_label] = df[["Episode", "Avg Reward"]].copy()
            series_fail[p_label] = df[["Episode", "normal_AVG_Failure"]].copy()
        except Exception as e:
            print(f"[missing_data] Error reading {global_xlsx}: {e}")

    if not series_reward:
        print("[missing_data] No usable data.")
        return

    wb = Workbook()
    wb.remove(wb.active)

    # -------- Sheet 1: Avg Reward over Episodes for all p --------
    _write_multi_p_line_sheet(
        wb=wb,
        sheet_name="AvgReward_all_p",
        series_dict=series_reward,
        y_title="Avg Reward",
        chart_title="Avg Reward over Episodes (missing_data)",
    )

    # -------- Sheet 2: Failure over Episodes for all p --------
    _write_multi_p_line_sheet(
        wb=wb,
        sheet_name="Failure_all_p",
        series_dict=series_fail,
        y_title="Avg Task Failure Rate (%)",
        chart_title="Avg Task Failure Rate (%) over Episodes (missing_data)",
    )

    out_path = os.path.join(out_dir, f"Final_Result_missing_data_state_{params.FAILURE_STATE}.xlsx")
    wb.save(out_path)
    print(f"[missing_data] Saved: {out_path}")


def _write_multi_p_line_sheet(wb, sheet_name: str, series_dict: dict, y_title: str, chart_title: str):

    ws = wb.create_sheet(sheet_name)

    # merge on Episode
    merged = None
    for p_label, df in series_dict.items():
        df = df.copy()
        # df columns: Episode, <valuecol>
        value_col = [c for c in df.columns if c != "Episode"][0]
        df = df.rename(columns={value_col: p_label})
        merged = df if merged is None else pd.merge(merged, df, on="Episode", how="outer")

    if merged is None or merged.empty:
        ws.append(["No data"])
        return

    _write_df_to_sheet(ws, merged)
    max_row = ws.max_row
    max_col = ws.max_column

    # chart
    chart = LineChart()
    chart.title = chart_title
    chart.x_axis.title = "Episode"
    chart.y_axis.title = y_title
    chart.width = 22
    chart.height = 10
    chart.legend.position = "r"

    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.set_categories(cats)

    # add all series columns (2..max_col)
    color_idx = 0
    for col in range(2, max_col + 1):
        data_ref = Reference(ws, min_col=col, max_col=col, min_row=1, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        s = chart.series[-1]
        color = DISTINCT_COLORS[color_idx % len(DISTINCT_COLORS)]
        color_idx += 1
        s.graphicalProperties.line.solidFill = color

    ws.add_chart(chart, "L2")


# -----------------------------
# trajectory_noise: bar chart mean±std on last 40 of normal_AVG_Failure
# -----------------------------
def run_trajectory_noise(root: str, out_dir: str):

    folders = _list_subfolders(root)
    dqn_folders = []
    for f in folders:
        if f.startswith("dqn_") and (not f.startswith("dqn_flat")):
            dqn_folders.append(f)

    if not dqn_folders:
        print("[trajectory_noise] No dqn_* folders found.")
        return

    rows = []  # (p_label, mean, std)
    for folder in sorted(dqn_folders):
        token = folder[len("dqn_"):]
        p_float = _safe_float_from_folder_token(token)
        p_label = f"p={p_float:.2f}"

        global_xlsx = os.path.join(root, folder, f"Global_state_{params.FAILURE_STATE}_dqn.xlsx")
        if not os.path.exists(global_xlsx):
            print(f"[trajectory_noise] Missing global file: {global_xlsx}")
            continue

        try:
            summ = pd.read_excel(global_xlsx, sheet_name="Summary")
            if "normal_AVG_Failure" not in summ.columns:
                raise KeyError(f"'normal_AVG_Failure' not found: {global_xlsx}")
            x = summ["normal_AVG_Failure"].dropna().values
            if len(x) == 0:
                continue
            last = x[-40:] if len(x) >= 40 else x
            mean = float(np.mean(last))
            std = float(np.std(last, ddof=0))  # population std
            rows.append((p_label, mean, std))
        except Exception as e:
            print(f"[trajectory_noise] Error reading {global_xlsx}: {e}")

    if not rows:
        print("[trajectory_noise] No usable data.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Failure_last40"

    ws.append(["p", "Mean(normal_AVG_Failure)", "Std(last40)"])
    for r in rows:
        ws.append(list(r))

    max_row = ws.max_row

    # Bar chart: Mean
    chart = BarChart()
    chart.type = "col"
    chart.title = "Failure Rate (normal_AVG_Failure) - Last 40 Episodes"
    chart.y_axis.title = "Mean"
    chart.x_axis.title = "p"
    chart.width = 22
    chart.height = 10
    chart.legend = None

    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data_mean = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    chart.add_data(data_mean, titles_from_data=True)
    chart.set_categories(cats)

    # ---- Error bars: custom = Std column ----
    # std range (only data rows)
    std_ref = Reference(ws, min_col=3, min_row=2, max_row=max_row)
    # openpyxl needs NumDataSource not Reference
    std_numref = NumRef(f"'{ws.title}'!$C$2:$C${max_row}")
    std_numdatasource = NumDataSource(numRef=std_numref)

    # apply to the first (and only) series
    if chart.series:
        ser = chart.series[0]
        eb = ErrorBars(errDir="y", errBarType="both", errValType="cust")
        eb.plus = std_numdatasource
        eb.minus = std_numdatasource
        ser.errBars = eb

        ser.graphicalProperties.solidFill = "1F77B4"

    ws.add_chart(chart, "E2")

    out_path = os.path.join(out_dir, f"Final_Result_trajectory_noise_state_{params.FAILURE_STATE}.xlsx")
    wb.save(out_path)
    print(f"[trajectory_noise] Saved: {out_path}")


# -----------------------------
# Main dispatcher
# -----------------------------
def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # <SCENARIO_TYPE>_results/<scenario>/
    scenarios_to_run = ["base", "missing_data", "trajectory_noise"]

    for scenario in scenarios_to_run:
        results_root = os.path.join(base_dir, f"{params.SCENARIO_TYPE}_results", scenario)
        _ensure_dir(results_root)

        if scenario == "base":
            base_root = os.path.join(base_dir, f"{params.SCENARIO_TYPE}_results", "base")
            if not os.path.isdir(base_root):
                print(f"[base] Folder not found, skipping: {base_root}")
                continue
            run_base_comparison(base_root=base_root, out_dir=results_root)

        elif scenario == "missing_data":
            md_root = os.path.join(base_dir, f"{params.SCENARIO_TYPE}_results", "missing_data")
            if not os.path.isdir(md_root):
                print(f"[missing_data] Folder not found, skipping: {md_root}")
                continue
            run_missing_data(root=md_root, out_dir=results_root)

        elif scenario == "trajectory_noise":
            tn_root = os.path.join(base_dir, f"{params.SCENARIO_TYPE}_results", "trajectory_noise")
            if not os.path.isdir(tn_root):
                print(f"[trajectory_noise] Folder not found, skipping: {tn_root}")
                continue
            run_trajectory_noise(root=tn_root, out_dir=results_root)

        else:
            print(f"[skip] Unknown scenario: {scenario}")


if __name__ == "__main__":
    main()
