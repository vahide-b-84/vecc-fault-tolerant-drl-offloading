# save.py
def save_params_and_logs(params, log_data_global, task_Assignments_info_global, rsu_logs_dict=None, rsu_assignments_dict=None):
    # Import standard libraries for file/path handling and data processing
    import os
    import pandas as pd
    import numpy as np

    # Detect flat mode:
    # - Flat_mode=True OR RSU-level logs are not passed => treat as flat (global-only) logging
    is_flat = bool(getattr(params, "Flat_mode", False)) or (rsu_logs_dict is None) or (rsu_assignments_dict is None)

    # Import OpenPyXL utilities for working with Excel files, sheets, and charts
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    # Excel hard limits (used to prevent writing invalid-sized sheets)
    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_COLS = 16_384

    def append_df_to_sheet(filename, df, sheet_name):
        # Append a DataFrame to an existing Excel sheet;
        # if the sheet does not exist, create it and write headers.
        book = load_workbook(filename)
        if sheet_name not in book.sheetnames:
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.close()
        else:
            sheet = book[sheet_name]
            # Append data rows only (no header) to avoid duplicating headers
            for r in dataframe_to_rows(df, index=False, header=False):
                sheet.append(r)
            book.save(filename)

    def _detect_episode_col(df, candidates=("episode", "Episode", "EPISODE")):
        # Detect the episode column name from common candidate options
        for c in candidates:
            if c in df.columns:
                return c
        raise KeyError(f"No episode column found. Tried {candidates}")

    def write_task_assignments_latest(filename: str, df, base_sheet_name="TaskAssignments", episode_col: str = None, episode_value: int = None):
        """
        Overwrite the 'TaskAssignments' sheet so it contains ONLY the latest episode records.
        - If episode_value is not provided, uses max(episode_col).
        - This prevents the TaskAssignments sheet from growing indefinitely.
        """
        if df is None or len(df) == 0:
            return

        ep_col = episode_col or _detect_episode_col(df)
        target_episode = episode_value if episode_value is not None else df[ep_col].max()
        df_latest = df[df[ep_col] == target_episode].copy()

        # Safety guards against writing an invalid Excel sheet size
        if df_latest.shape[1] > EXCEL_MAX_COLS:
            raise ValueError(f"Too many columns for Excel: {df_latest.shape[1]} > {EXCEL_MAX_COLS}")
        if df_latest.shape[0] + 1 > EXCEL_MAX_ROWS:
            raise ValueError(f"Too many rows for Excel: {df_latest.shape[0]} data rows")

        # Load or create workbook
        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            wb = Workbook()

        # Remove old TaskAssignments sheet if it exists (full replace)
        if base_sheet_name in wb.sheetnames:
            wb.remove(wb[base_sheet_name])
            # Ensure at least one sheet exists in a new workbook
            if not wb.sheetnames:
                wb.create_sheet(title="Sheet")

        # Create fresh TaskAssignments sheet and write header + rows
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(list(df_latest.columns))
        for row in dataframe_to_rows(df_latest, index=False, header=False):
            ws.append(row)

        wb.save(filename)
        wb.close()

    def add_charts_to_sheet(wb, sheet_name):
        # Add (or refresh) charts for a given sheet in the workbook
        ws = wb[sheet_name]
        # Remove old charts to avoid stacking duplicates on repeated saves
        ws._charts.clear()

        max_row = ws.max_row
        max_col = ws.max_column

        if sheet_name == 'Logs':
            # Chart 1: Reward per Episode (Avg Reward and Episode Reward)
            chart_rewards = LineChart()
            chart_rewards.title = "Reward per Episode"
            chart_rewards.y_axis.title = 'Reward'
            chart_rewards.x_axis.title = 'Episode'

            # Columns:
            # A: Episode, B: Avg Reward, C: Episode Reward, D: Avg Delay, E: Episode Delay
            data_rewards = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)

            chart_rewards.add_data(data_rewards, titles_from_data=True)
            chart_rewards.set_categories(categories)
            chart_rewards.width = 20
            chart_rewards.height = 10
            ws.add_chart(chart_rewards, f"{get_column_letter(max_col+2)}2")

            # Chart 2: Delay per Episode (Avg Delay and Episode Delay)
            chart_delays = LineChart()
            chart_delays.title = "Delay per Episode"
            chart_delays.y_axis.title = 'Delay'
            chart_delays.x_axis.title = 'Episode'

            data_delays = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=max_row)
            chart_delays.add_data(data_delays, titles_from_data=True)
            chart_delays.set_categories(categories)
            chart_delays.width = 20
            chart_delays.height = 10
            ws.add_chart(chart_delays, f"{get_column_letter(max_col+2)}20")

        elif sheet_name == 'Summary':
            # Summary sheet expected columns:
            # A: episode, B: Failure, C: Success, D: Total, E: FailureRate
            failure_col_idx = 5  # Column E = FailureRate
            episode_col_idx = 1  # Column A = episode
            avg_col_idx = 6      # Column F will be added for smoothed failure

            # Add headers for computed rolling average columns
            ws.cell(row=1, column=avg_col_idx, value='normal_AVG_Failure')
            ws.cell(row=1, column=avg_col_idx + 1, value='AVG_Failure')

            # Read FailureRate values from column E
            failure_values = [row[0] for row in ws.iter_rows(
                min_row=2, min_col=failure_col_idx, max_col=failure_col_idx, values_only=True
            )]

            # Compute moving average over last 40 episodes (or fewer at the beginning)
            avg_values = []
            for i in range(len(failure_values)):
                if i >= 39:
                    avg = np.mean(failure_values[i - 39:i + 1])
                else:
                    avg = np.mean(failure_values[:i + 1])
                avg_values.append(avg)

            # Write moving average values to columns F and G
            # Column F: normal_AVG_Failure (already a percentage-like value)
            # Column G: scaled AVG_Failure using Total/100 (project-specific interpretation)
            for i, val in enumerate(avg_values, start=2):
                ws.cell(row=i, column=avg_col_idx, value=val)
                ws.cell(row=i, column=avg_col_idx + 1, value=val * ws.cell(row=i, column=4).value / 100)

            # Plot the smoothed failure rate in a line chart
            line_chart = LineChart()
            line_chart.title = "normal AVG Failure Rate Over Episodes"
            line_chart.style = 6
            line_chart.x_axis.title = 'Episode'
            line_chart.y_axis.title = 'normal_AVG_Failure'
            line_chart.width = 20
            line_chart.height = 10
            line_chart.legend = None

            episodes = Reference(ws, min_col=episode_col_idx, min_row=2, max_row=len(failure_values) + 1)
            avg_failure_ref = Reference(ws, min_col=avg_col_idx, min_row=1, max_row=len(failure_values) + 1)

            line_chart.add_data(avg_failure_ref, titles_from_data=True)
            line_chart.set_categories(episodes)
            ws.add_chart(line_chart, f"{get_column_letter(ws.max_column + 2)}2")

    # Determine the directory of the current script file (base path for relative files)
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Determine scenario tag used in results directory naming
    scenario = getattr(params, "scenario", "base")

    # Determine probability parameter p based on scenario type (used in folder naming)
    if scenario == "trajectory_noise":
        p = getattr(params, "trajectory_noise_p", 0.0)
    elif scenario == "missing_data":
        p = getattr(params, "missing_data_p", 0.0)
    elif scenario == "base":
        p = 0.0
    else:
        p = 0.0

    # Build model tag used in folder naming (e.g., "dqn_0_30" or "ppo_flat_0_00")
    flat_suffix = "_flat" if is_flat else ""
    model_tag = f"{params.model_summary}{flat_suffix}_{p:.2f}".replace(".", "_")

    # Create results directory path: <SCENARIO_TYPE>_results/<scenario>/<model_tag>
    results_dir = os.path.join(current_dir, f"{params.SCENARIO_TYPE}_results", scenario, model_tag)
    os.makedirs(results_dir, exist_ok=True)

    # Load server info from the appropriate Excel file (homogeneous vs heterogeneous)
    fname = os.path.join(current_dir, 'homogeneous_server_info.xlsx') if params.SCENARIO_TYPE == 'homogeneous' else os.path.join(current_dir, 'heterogeneous_server_info.xlsx')

    # Select correct sheet by FAILURE_STATE (low/med/high)
    sheet_name = f'{params.SCENARIO_TYPE.capitalize()}_state_{params.FAILURE_STATE}'
    server_info = pd.read_excel(fname, sheet_name=sheet_name)

    # Load task parameters and server info into DataFrames (for saving alongside results)
    task_df = pd.read_excel(os.path.join(current_dir, 'task_parameters.xlsx'))
    df_tasks = pd.DataFrame(task_df)
    df_servers = pd.DataFrame(server_info)

    # Define the global output Excel filename
    global_filename = os.path.join(results_dir, f"Global_state_{params.FAILURE_STATE}_{params.model_summary}.xlsx")

    # If global file exists, read existing episodes from Logs to prevent duplicates
    if os.path.exists(global_filename):
        existing_global = pd.read_excel(global_filename, sheet_name='Logs')
        existing_episodes = set(existing_global['Episode'].tolist())
    else:
        existing_episodes = set()

    # Keep only logs/assignments whose episode is not already present in the existing file
    new_logs = [log for log in log_data_global if log[0] not in existing_episodes]
    new_assignments = [a for a in task_Assignments_info_global if a[0] not in existing_episodes]

    # Convert new global logs to a DataFrame with fixed column mapping
    df_new_logs = pd.DataFrame([{
        'Episode': log[0],
        'Avg Reward': log[1],
        'Episode Reward': log[2],
        'Avg Delay': log[3],
        'Episode Delay': log[4]
    } for log in new_logs])

    # Convert new assignment records to DataFrame with flat/non-flat schemas
    if is_flat:
        # Flat schema: includes RSU selection + (primary, backup, z) details
        df_new_assignments = pd.DataFrame(new_assignments, columns=[
            'episode', 'task_id', 'vehicle_id',
            'original_rsu', 'submitted_time',
            'selected_RSU', 'start_executing', 'final_rsu', 'finished_time',
            'primary', 'primary_start', 'primary_end', 'primary_status',
            'backup', 'backup_start', 'backup_end', 'backup_status',
            'z', 'executaion_status', 'deadline_flag', 'final_status'
        ])
    else:
        # Two-level schema: global sheet stores only global decisions/outcomes (local details are in per-RSU files)
        df_new_assignments = pd.DataFrame(new_assignments, columns=[
            'episode', 'task_id', 'vehicle_id', 'original_rsu', 'submitted_time',
            'selected_RSU', 'start_executing', 'final_rsu', 'finished_time',
            'executaion_status', 'deadline_flag', 'final_status'
        ])

    # Build summary table: count Success/Failure per episode based on final_status
    summary_df = (
        df_new_assignments.groupby(['episode', 'final_status'])
        .size().unstack(fill_value=0)
        .rename(columns={'f': 'Failure', 's': 'Success'})
        .reset_index()
    )
    summary_df = summary_df.reindex(columns=['episode', 'Failure', 'Success'], fill_value=0)
    summary_df['Total'] = summary_df['Success'] + summary_df['Failure']
    summary_df['FailureRate'] = 100 * summary_df['Failure'] / summary_df['Total'].replace(0, 1)
    summary_df = summary_df.fillna(0)

    if not os.path.exists(global_filename):
        # Create Params DataFrame from params object's attributes (snapshot of configuration)
        df_params = pd.DataFrame(list(vars(params).items()), columns=['Parameter', 'Value'])

        # Only keep the latest episode in TaskAssignments on initial creation
        ep_col = _detect_episode_col(df_new_assignments)
        last_ep = df_new_assignments[ep_col].max()
        df_last_assignments = df_new_assignments[df_new_assignments[ep_col] == last_ep].copy()

        # Write all sheets into a new Excel file (first creation)
        with pd.ExcelWriter(global_filename) as writer:
            df_params.to_excel(writer, sheet_name='Params', index=False)
            df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
            df_servers.to_excel(writer, sheet_name='Servers', index=False)
            df_new_logs.to_excel(writer, sheet_name='Logs', index=False)
            df_last_assignments.to_excel(writer, sheet_name='TaskAssignments', index=False)  # only latest episode
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

    else:
        # Append new logs to Logs sheet (keep growing over episodes)
        append_df_to_sheet(global_filename, df_new_logs, 'Logs')

        # Replace TaskAssignments with latest-episode-only records (prevents sheet growth)
        write_task_assignments_latest(global_filename, df_new_assignments, base_sheet_name='TaskAssignments')

        # Append summary rows to Summary sheet
        append_df_to_sheet(global_filename, summary_df, 'Summary')

    # Reload workbook to add/refresh charts on Logs and Summary sheets
    wb = load_workbook(global_filename)
    if 'Logs' in wb.sheetnames:
        add_charts_to_sheet(wb, 'Logs')
    if 'Summary' in wb.sheetnames:
        add_charts_to_sheet(wb, 'Summary')
    wb.save(global_filename)
    print("Global log updated.")

    # Iterate through each RSU and write/update its own Excel file (only in non-flat mode)
    if not is_flat:
        # Cloud servers are shared; appended to each RSU's server list for completeness
        cloud_servers = df_servers[df_servers['Server_Type'] == 'Cloud']

        for rsu_id in rsu_logs_dict.keys():
            filename_rsu = os.path.join(results_dir, f"{rsu_id}_state_{params.FAILURE_STATE}_{params.model_summary}.xlsx")

            local_logs = rsu_logs_dict[rsu_id]
            local_assignments = rsu_assignments_dict[rsu_id]

            # Prevent duplicate episodes per RSU file as well
            if os.path.exists(filename_rsu):
                existing_rsu = pd.read_excel(filename_rsu, sheet_name='Logs')
                existing_episodes_rsu = set(existing_rsu['Episode'].tolist())
            else:
                existing_episodes_rsu = set()

            new_logs_rsu = [log for log in local_logs if log[0] not in existing_episodes_rsu]
            new_assignments_rsu = [a for a in local_assignments if a[0] not in existing_episodes_rsu]

            # RSU Logs dataframe
            df_logs_rsu = pd.DataFrame([{
                'Episode': log[0],
                'Avg Reward': log[1],
                'Episode Reward': log[2],
                'Avg Delay': log[3],
                'Episode Delay': log[4]
            } for log in new_logs_rsu])

            # RSU TaskAssignments dataframe (local decisions: primary/backup/z + status)
            df_assignments_rsu = pd.DataFrame(new_assignments_rsu, columns=[
                'episode', 'task_id', 'vehicle_id', 'Primary', 'Primary_Start', 'Primary_End', 'Primary_Status',
                'Backup', 'Backup_Start', 'Backup_End', 'Backup_Status', 'Z', 'executaion_status'
            ])

            # Per-RSU server list: (edge servers for that RSU) + (all cloud servers)
            rsu_servers = pd.concat([df_servers[df_servers['RSU_ID'] == rsu_id], cloud_servers], ignore_index=True)

            # Summary per RSU: success/failure counts over 'executaion_status'
            summary_rsu = (
                df_assignments_rsu.groupby(['episode', 'executaion_status'])
                .size().unstack(fill_value=0)
                .rename(columns={'f': 'Failure', 's': 'Success'})
                .reset_index()
            )
            summary_rsu = summary_rsu.reindex(columns=['episode', 'Failure', 'Success'], fill_value=0)
            summary_rsu['Total'] = summary_rsu['Success'] + summary_rsu['Failure']
            summary_rsu['FailureRate'] = 100 * summary_rsu['Failure'] / summary_rsu['Total'].replace(0, 1)
            summary_rsu = summary_rsu.fillna(0)

            if not os.path.exists(filename_rsu):
                # On first creation, keep only the latest episode in TaskAssignments
                ep_col_rsu = _detect_episode_col(df_assignments_rsu) if len(df_assignments_rsu) else 'episode'
                last_ep_rsu = df_assignments_rsu[ep_col_rsu].max() if len(df_assignments_rsu) else None
                df_last_assignments_rsu = (
                    df_assignments_rsu[df_assignments_rsu[ep_col_rsu] == last_ep_rsu].copy()
                    if last_ep_rsu is not None else df_assignments_rsu.copy()
                )

                with pd.ExcelWriter(filename_rsu) as writer:
                    rsu_servers.to_excel(writer, sheet_name='Servers', index=False)
                    df_logs_rsu.to_excel(writer, sheet_name='Logs', index=False)
                    df_last_assignments_rsu.to_excel(writer, sheet_name='TaskAssignments', index=False)
                    summary_rsu.to_excel(writer, sheet_name='Summary', index=False)
            else:
                # Append Logs and Summary; overwrite TaskAssignments with latest episode only
                append_df_to_sheet(filename_rsu, df_logs_rsu, 'Logs')
                write_task_assignments_latest(filename_rsu, df_assignments_rsu, base_sheet_name='TaskAssignments')
                append_df_to_sheet(filename_rsu, summary_rsu, 'Summary')

            # Refresh charts for RSU file as well
            wb_rsu = load_workbook(filename_rsu)
            if 'Logs' in wb_rsu.sheetnames:
                add_charts_to_sheet(wb_rsu, 'Logs')
            if 'Summary' in wb_rsu.sheetnames:
                add_charts_to_sheet(wb_rsu, 'Summary')
            wb_rsu.save(filename_rsu)
            print(f"{rsu_id} log updated.")

    # ------------------------------------------------------------
    # REPLACE-COPY base (p=0) bi-level DQN results for other scenarios
    # base/dqn_0_00  ->  missing_data/dqn_0_00  &  trajectory_noise/dqn_0_00
    # ------------------------------------------------------------
    if (not is_flat) and (scenario == "base") and (params.model_summary == "dqn") and (abs(p - 0.0) < 1e-12):
        import shutil

        # src_dir: .../<SCENARIO_TYPE>_results/base/dqn_0_00
        src_dir = results_dir
        base_root = os.path.join(current_dir, f"{params.SCENARIO_TYPE}_results")

        for dst_scenario in ["missing_data", "trajectory_noise"]:
            dst_dir = os.path.join(base_root, dst_scenario, os.path.basename(src_dir))

            # 
            if os.path.exists(dst_dir):
                shutil.rmtree(dst_dir)

            # 
            os.makedirs(os.path.dirname(dst_dir), exist_ok=True)
            shutil.copytree(src_dir, dst_dir)

            print(f"[REPLACED COPY] {src_dir}  ->  {dst_dir}")


    return results_dir
